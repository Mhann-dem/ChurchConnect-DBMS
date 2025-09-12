# backend/churchconnect/reports/services.py

import os
import csv
import json
import logging
import tempfile
from pathlib import Path
from typing import Dict, List, Any, Optional, Union
from io import StringIO, BytesIO
from datetime import datetime, timedelta
from contextlib import contextmanager

import openpyxl
from django.conf import settings
from django.db.models import Q, Count, Sum, Avg
from django.utils import timezone
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.core.files.storage import default_storage
from django.core.exceptions import ValidationError

from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch

from .models import Report, ReportRun, ReportTemplate

# Import your app models
try:
    from members.models import Member
    from pledges.models import Pledge
    from groups.models import Group
    from families.models import Family
except ImportError:
    # Handle missing models gracefully
    Member = Pledge = Group = Family = None

logger = logging.getLogger(__name__)


class ReportGeneratorError(Exception):
    """Custom exception for report generation errors"""
    pass


class ReportGeneratorService:
    """Service class for generating various types of reports"""
    
    def __init__(self):
        self.media_root = Path(settings.MEDIA_ROOT)
        self.reports_dir = self.media_root / 'reports'
        
        # Ensure reports directory exists
        self.reports_dir.mkdir(parents=True, exist_ok=True)
        
        # Maximum file size (100MB)
        self.max_file_size = 100 * 1024 * 1024
        
        # Default timeout (5 minutes)
        self.generation_timeout = 300
    
    @contextmanager
    def _safe_file_creation(self, file_path: Path):
        """Context manager for safe file creation with cleanup"""
        temp_path = None
        try:
            # Create temporary file first
            temp_path = file_path.with_suffix(file_path.suffix + '.tmp')
            yield temp_path
            
            # Move to final location if successful
            if temp_path.exists():
                temp_path.rename(file_path)
                
        except Exception:
            # Clean up temporary file on error
            if temp_path and temp_path.exists():
                temp_path.unlink()
            raise
    
    def generate_report(self, report: Report, report_run: ReportRun) -> Dict[str, Any]:
        """Generate a report based on configuration"""
        try:
            logger.info(f"Starting report generation for report {report.id}")
            
            # Validate inputs
            self._validate_report_config(report)
            
            # Get data based on report type
            data = self._get_report_data(
                report.report_type,
                report.filters or {},
                report.columns or []
            )
            
            logger.info(f"Retrieved {len(data)} records for report {report.id}")
            
            # Generate file based on format
            file_path = self._generate_file(
                data,
                report.format,
                report.name,
                report_run.id
            )
            
            # Get file size and validate
            file_size = file_path.stat().st_size
            if file_size > self.max_file_size:
                raise ReportGeneratorError(f"Generated file exceeds maximum size limit ({self.max_file_size} bytes)")
            
            logger.info(f"Report {report.id} generated successfully: {file_path}")
            
            return {
                'success': True,
                'file_path': str(file_path),
                'file_size': file_size,
                'record_count': len(data)
            }
            
        except Exception as e:
            logger.error(f"Error generating report {report.id}: {str(e)}", exc_info=True)
            return {
                'success': False,
                'error': str(e)
            }
    
    def generate_adhoc_report(self, report_config: Dict[str, Any], report_run: ReportRun) -> Dict[str, Any]:
        """Generate an ad-hoc report without saving configuration"""
        try:
            logger.info(f"Starting ad-hoc report generation for run {report_run.id}")
            
            # Validate configuration
            self._validate_adhoc_config(report_config)
            
            data = self._get_report_data(
                report_config['report_type'],
                report_config.get('filters', {}),
                report_config.get('columns', [])
            )
            
            name = report_config.get('name', f"adhoc_{report_config['report_type']}")
            
            file_path = self._generate_file(
                data,
                report_config['format'],
                name,
                report_run.id
            )
            
            file_size = file_path.stat().st_size
            if file_size > self.max_file_size:
                raise ReportGeneratorError(f"Generated file exceeds maximum size limit ({self.max_file_size} bytes)")
            
            logger.info(f"Ad-hoc report generated successfully: {file_path}")
            
            return {
                'success': True,
                'file_path': str(file_path),
                'file_size': file_size,
                'record_count': len(data)
            }
            
        except Exception as e:
            logger.error(f"Error generating ad-hoc report for run {report_run.id}: {str(e)}", exc_info=True)
            return {
                'success': False,
                'error': str(e)
            }
    
    def _validate_report_config(self, report: Report) -> None:
        """Validate report configuration"""
        if not report.report_type:
            raise ReportGeneratorError("Report type is required")
        
        if report.report_type not in dict(Report.REPORT_TYPE_CHOICES):
            raise ReportGeneratorError(f"Invalid report type: {report.report_type}")
        
        if not report.format:
            raise ReportGeneratorError("Report format is required")
        
        if report.format not in dict(Report.FORMAT_CHOICES):
            raise ReportGeneratorError(f"Invalid format: {report.format}")
    
    def _validate_adhoc_config(self, config: Dict[str, Any]) -> None:
        """Validate ad-hoc report configuration"""
        if 'report_type' not in config:
            raise ReportGeneratorError("Report type is required")
        
        if config['report_type'] not in dict(Report.REPORT_TYPE_CHOICES):
            raise ReportGeneratorError(f"Invalid report type: {config['report_type']}")
        
        if 'format' not in config:
            raise ReportGeneratorError("Format is required")
        
        if config['format'] not in dict(Report.FORMAT_CHOICES):
            raise ReportGeneratorError(f"Invalid format: {config['format']}")
    
    def _get_report_data(self, report_type: str, filters: Dict[str, Any], columns: List[str]) -> List[Dict[str, Any]]:
        """Get data for specific report type"""
        try:
            if report_type == 'members':
                return self._get_members_data(filters, columns)
            elif report_type == 'pledges':
                return self._get_pledges_data(filters, columns)
            elif report_type == 'groups':
                return self._get_groups_data(filters, columns)
            elif report_type == 'families':
                return self._get_families_data(filters, columns)
            elif report_type == 'statistics':
                return self._get_statistics_data(filters, columns)
            else:
                raise ReportGeneratorError(f"Unknown report type: {report_type}")
        except Exception as e:
            logger.error(f"Error retrieving data for report type {report_type}: {str(e)}")
            raise ReportGeneratorError(f"Failed to retrieve data: {str(e)}")
    
    def _get_members_data(self, filters: Dict[str, Any], columns: List[str]) -> List[Dict[str, Any]]:
        """Get members data with filters"""
        if Member is None:
            raise ReportGeneratorError("Members model not available")
        
        try:
            queryset = Member.objects.select_related('family').prefetch_related('groups', 'pledges')
            
            # Apply filters
            if filters.get('age_min'):
                birth_date_max = timezone.now().date() - timedelta(days=filters['age_min'] * 365)
                queryset = queryset.filter(date_of_birth__lte=birth_date_max)
            
            if filters.get('age_max'):
                birth_date_min = timezone.now().date() - timedelta(days=filters['age_max'] * 365)
                queryset = queryset.filter(date_of_birth__gte=birth_date_min)
            
            if filters.get('gender'):
                queryset = queryset.filter(gender=filters['gender'])
            
            if filters.get('groups'):
                queryset = queryset.filter(groups__id__in=filters['groups']).distinct()
            
            if filters.get('has_pledged') is not None:
                if filters['has_pledged']:
                    queryset = queryset.filter(pledges__isnull=False).distinct()
                else:
                    queryset = queryset.filter(pledges__isnull=True)
            
            if filters.get('registration_date_from'):
                queryset = queryset.filter(registration_date__gte=filters['registration_date_from'])
            
            if filters.get('registration_date_to'):
                queryset = queryset.filter(registration_date__lte=filters['registration_date_to'])
            
            if filters.get('is_active') is not None:
                queryset = queryset.filter(is_active=filters['is_active'])
            
            # Limit queryset size to prevent memory issues
            queryset = queryset[:10000]
            
            # Default columns if none specified
            if not columns:
                columns = [
                    'first_name', 'last_name', 'email', 'phone', 'date_of_birth',
                    'gender', 'address', 'registration_date', 'preferred_contact_method'
                ]
            
            # Convert queryset to list of dictionaries
            data = []
            for member in queryset:
                row = {}
                for column in columns:
                    try:
                        if hasattr(member, column):
                            value = getattr(member, column)
                            if isinstance(value, datetime):
                                value = value.strftime('%Y-%m-%d %H:%M:%S')
                            elif value is None:
                                value = ''
                            row[column] = value
                        elif column == 'age':
                            if member.date_of_birth:
                                age = (timezone.now().date() - member.date_of_birth).days // 365
                                row[column] = age
                            else:
                                row[column] = ''
                        elif column == 'groups':
                            row[column] = ', '.join([group.name for group in member.groups.all()])
                        elif column == 'family_name':
                            row[column] = member.family.family_name if member.family else ''
                        elif column == 'total_pledged':
                            total = member.pledges.aggregate(Sum('amount'))['amount__sum'] or 0
                            row[column] = total
                        else:
                            row[column] = ''
                    except Exception as e:
                        logger.warning(f"Error processing column {column} for member {member.id}: {str(e)}")
                        row[column] = ''
                data.append(row)
            
            return data
            
        except Exception as e:
            logger.error(f"Error in _get_members_data: {str(e)}")
            raise ReportGeneratorError(f"Failed to retrieve members data: {str(e)}")
    
    def _get_pledges_data(self, filters: Dict[str, Any], columns: List[str]) -> List[Dict[str, Any]]:
        """Get pledges data with filters"""
        if Pledge is None:
            raise ReportGeneratorError("Pledges model not available")
        
        try:
            queryset = Pledge.objects.select_related('member')
            
            # Apply filters
            if filters.get('amount_min'):
                queryset = queryset.filter(amount__gte=filters['amount_min'])
            
            if filters.get('amount_max'):
                queryset = queryset.filter(amount__lte=filters['amount_max'])
            
            if filters.get('frequency'):
                queryset = queryset.filter(frequency=filters['frequency'])
            
            if filters.get('status'):
                queryset = queryset.filter(status=filters['status'])
            
            if filters.get('start_date_from'):
                queryset = queryset.filter(start_date__gte=filters['start_date_from'])
            
            if filters.get('start_date_to'):
                queryset = queryset.filter(start_date__lte=filters['start_date_to'])
            
            # Limit queryset size
            queryset = queryset[:10000]
            
            # Default columns
            if not columns:
                columns = [
                    'member_name', 'member_email', 'amount', 'frequency',
                    'start_date', 'end_date', 'status', 'created_at'
                ]
            
            data = []
            for pledge in queryset:
                row = {}
                for column in columns:
                    try:
                        if hasattr(pledge, column):
                            value = getattr(pledge, column)
                            if isinstance(value, datetime):
                                value = value.strftime('%Y-%m-%d %H:%M:%S')
                            elif value is None:
                                value = ''
                            row[column] = value
                        elif column == 'member_name':
                            row[column] = f"{pledge.member.first_name} {pledge.member.last_name}"
                        elif column == 'member_email':
                            row[column] = pledge.member.email
                        elif column == 'member_phone':
                            row[column] = pledge.member.phone
                        else:
                            row[column] = ''
                    except Exception as e:
                        logger.warning(f"Error processing column {column} for pledge {pledge.id}: {str(e)}")
                        row[column] = ''
                data.append(row)
            
            return data
            
        except Exception as e:
            logger.error(f"Error in _get_pledges_data: {str(e)}")
            raise ReportGeneratorError(f"Failed to retrieve pledges data: {str(e)}")
    
    def _get_groups_data(self, filters: Dict[str, Any], columns: List[str]) -> List[Dict[str, Any]]:
        """Get groups data with filters"""
        if Group is None:
            raise ReportGeneratorError("Groups model not available")
        
        try:
            queryset = Group.objects.prefetch_related('members')
            
            # Apply filters
            if filters.get('active') is not None:
                queryset = queryset.filter(active=filters['active'])
            
            if filters.get('min_members'):
                queryset = queryset.annotate(
                    member_count=Count('members')
                ).filter(member_count__gte=filters['min_members'])
            
            # Limit queryset size
            queryset = queryset[:1000]
            
            # Default columns
            if not columns:
                columns = [
                    'name', 'description', 'leader_name', 'meeting_schedule',
                    'member_count', 'active', 'created_at'
                ]
            
            data = []
            for group in queryset:
                row = {}
                for column in columns:
                    try:
                        if hasattr(group, column):
                            value = getattr(group, column)
                            if isinstance(value, datetime):
                                value = value.strftime('%Y-%m-%d %H:%M:%S')
                            elif value is None:
                                value = ''
                            row[column] = value
                        elif column == 'member_count':
                            row[column] = group.members.count()
                        else:
                            row[column] = ''
                    except Exception as e:
                        logger.warning(f"Error processing column {column} for group {group.id}: {str(e)}")
                        row[column] = ''
                data.append(row)
            
            return data
            
        except Exception as e:
            logger.error(f"Error in _get_groups_data: {str(e)}")
            raise ReportGeneratorError(f"Failed to retrieve groups data: {str(e)}")
    
    def _get_families_data(self, filters: Dict[str, Any], columns: List[str]) -> List[Dict[str, Any]]:
        """Get families data with filters"""
        if Family is None:
            raise ReportGeneratorError("Families model not available")
        
        try:
            queryset = Family.objects.prefetch_related('members')
            
            # Apply filters
            if filters.get('min_members'):
                queryset = queryset.annotate(
                    member_count=Count('members')
                ).filter(member_count__gte=filters['min_members'])
            
            # Limit queryset size
            queryset = queryset[:1000]
            
            # Default columns
            if not columns:
                columns = [
                    'family_name', 'primary_contact_name', 'primary_contact_email',
                    'address', 'member_count', 'created_at'
                ]
            
            data = []
            for family in queryset:
                row = {}
                for column in columns:
                    try:
                        if hasattr(family, column):
                            value = getattr(family, column)
                            if isinstance(value, datetime):
                                value = value.strftime('%Y-%m-%d %H:%M:%S')
                            elif value is None:
                                value = ''
                            row[column] = value
                        elif column == 'member_count':
                            row[column] = family.members.count()
                        elif column == 'primary_contact_name':
                            if family.primary_contact:
                                row[column] = f"{family.primary_contact.first_name} {family.primary_contact.last_name}"
                            else:
                                row[column] = ''
                        elif column == 'primary_contact_email':
                            row[column] = family.primary_contact.email if family.primary_contact else ''
                        else:
                            row[column] = ''
                    except Exception as e:
                        logger.warning(f"Error processing column {column} for family {family.id}: {str(e)}")
                        row[column] = ''
                data.append(row)
            
            return data
            
        except Exception as e:
            logger.error(f"Error in _get_families_data: {str(e)}")
            raise ReportGeneratorError(f"Failed to retrieve families data: {str(e)}")
    
    def _get_statistics_data(self, filters: Dict[str, Any], columns: List[str]) -> List[Dict[str, Any]]:
        """Get statistics data"""
        try:
            data = []
            
            # Members statistics
            if Member:
                total_members = Member.objects.count()
                active_members = Member.objects.filter(is_active=True).count()
                new_members_this_month = Member.objects.filter(
                    registration_date__gte=timezone.now().replace(day=1)
                ).count()
                
                data.extend([
                    {'category': 'Members', 'metric': 'Total Members', 'value': total_members},
                    {'category': 'Members', 'metric': 'Active Members', 'value': active_members},
                    {'category': 'Members', 'metric': 'New Members This Month', 'value': new_members_this_month},
                ])
                
                # Gender distribution
                gender_stats = Member.objects.values('gender').annotate(count=Count('id'))
                for stat in gender_stats:
                    if stat['gender']:
                        data.append({
                            'category': 'Demographics',
                            'metric': f"Members - {stat['gender'].title()}",
                            'value': stat['count']
                        })
            
            # Pledges statistics
            if Pledge:
                total_pledges = Pledge.objects.count()
                active_pledges = Pledge.objects.filter(status='active').count()
                total_pledge_amount = Pledge.objects.aggregate(Sum('amount'))['amount__sum'] or 0
                
                data.extend([
                    {'category': 'Pledges', 'metric': 'Total Pledges', 'value': total_pledges},
                    {'category': 'Pledges', 'metric': 'Active Pledges', 'value': active_pledges},
                    {'category': 'Pledges', 'metric': 'Total Pledge Amount', 'value': total_pledge_amount},
                ])
            
            # Groups statistics
            if Group:
                total_groups = Group.objects.count()
                active_groups = Group.objects.filter(active=True).count()
                
                data.extend([
                    {'category': 'Groups', 'metric': 'Total Groups', 'value': total_groups},
                    {'category': 'Groups', 'metric': 'Active Groups', 'value': active_groups},
                ])
            
            return data
            
        except Exception as e:
            logger.error(f"Error in _get_statistics_data: {str(e)}")
            raise ReportGeneratorError(f"Failed to retrieve statistics data: {str(e)}")
    
    def _generate_file(self, data: List[Dict[str, Any]], format_type: str, name: str, run_id: str) -> Path:
        """Generate file in specified format"""
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        safe_name = "".join(c for c in name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        filename = f"{safe_name}_{timestamp}_{run_id}"
        
        if format_type == 'csv':
            return self._generate_csv(data, filename)
        elif format_type == 'excel':
            return self._generate_excel(data, filename)
        elif format_type == 'pdf':
            return self._generate_pdf(data, filename, name)
        elif format_type == 'json':
            return self._generate_json(data, filename)
        else:
            raise ReportGeneratorError(f"Unsupported format: {format_type}")
    
    def _generate_csv(self, data: List[Dict[str, Any]], filename: str) -> Path:
        """Generate CSV file"""
        file_path = self.reports_dir / f"{filename}.csv"
        
        with self._safe_file_creation(file_path) as temp_path:
            with open(temp_path, 'w', newline='', encoding='utf-8') as csvfile:
                if not data:
                    csvfile.write("No data available")
                else:
                    fieldnames = data[0].keys()
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(data)
        
        return file_path
    
    def _generate_excel(self, data: List[Dict[str, Any]], filename: str) -> Path:
        """Generate Excel file"""
        file_path = self.reports_dir / f"{filename}.xlsx"
        
        with self._safe_file_creation(file_path) as temp_path:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Report Data"
            
            if not data:
                worksheet.cell(row=1, column=1, value="No data available")
            else:
                # Write headers
                headers = list(data[0].keys())
                for col, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=1, column=col, value=header)
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                # Write data
                for row_idx, row_data in enumerate(data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        value = row_data.get(header, '')
                        # Handle large numbers and dates properly
                        if isinstance(value, (int, float)):
                            worksheet.cell(row=row_idx, column=col_idx, value=value)
                        else:
                            worksheet.cell(row=row_idx, column=col_idx, value=str(value))
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(temp_path)
        
        return file_path
    
    def _generate_pdf(self, data: List[Dict[str, Any]], filename: str, title: str) -> Path:
        """Generate PDF file"""
        file_path = self.reports_dir / f"{filename}.pdf"
        
        with self._safe_file_creation(file_path) as temp_path:
            doc = SimpleDocTemplate(str(temp_path), pagesize=A4)
            story = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            
            # Title
            story.append(Paragraph(title, title_style))
            story.append(Spacer(1, 12))
            
            # Generation info
            generation_info = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
            story.append(Paragraph(generation_info, styles['Normal']))
            story.append(Spacer(1, 12))
            
            if not data:
                story.append(Paragraph("No data available", styles['Normal']))
            else:
                # Create table with limited columns for better fit
                headers = list(data[0].keys())[:6]  # Limit to 6 columns
                table_data = [headers]
                
                for row in data[:100]:  # Limit to 100 rows for PDF
                    table_data.append([str(row.get(header, ''))[:50] for header in headers])
                
                if len(data) > 100:
                    story.append(Paragraph(f"Note: Showing first 100 rows of {len(data)} total records", styles['Italic']))
                    story.append(Spacer(1, 12))
                
                if len(data[0].keys()) > 6:
                    story.append(Paragraph("Note: Only first 6 columns shown due to page width constraints", styles['Italic']))
                    story.append(Spacer(1, 12))
                
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                
                story.append(table)
            
            # Record count
            record_count = f"Total Records: {len(data)}"
            story.append(Spacer(1, 12))
            story.append(Paragraph(record_count, styles['Normal']))
            
            doc.build(story)
        
        return file_path
    
    def _generate_json(self, data: List[Dict[str, Any]], filename: str) -> Path:
        """Generate JSON file"""
        file_path = self.reports_dir / f"{filename}.json"
        
        with self._safe_file_creation(file_path) as temp_path:
            report_data = {
                'generated_at': timezone.now().isoformat(),
                'record_count': len(data),
                'data': data
            }
            
            with open(temp_path, 'w', encoding='utf-8') as jsonfile:
                json.dump(report_data, jsonfile, indent=2, ensure_ascii=False, default=str)
        
        return file_path


class ReportSchedulerService:
    """Service for managing scheduled reports"""
    
    def __init__(self):
        self.generator = ReportGeneratorService()
        self.max_concurrent_reports = 5
    
    def run_scheduled_reports(self) -> List[Dict[str, Any]]:
        """Run all scheduled reports that are due"""
        now = timezone.now()
        due_reports = Report.objects.filter(
            is_scheduled=True,
            is_active=True,
            next_run__lte=now
        ).select_related('created_by')[:self.max_concurrent_reports]
        
        results = []
        
        for report in due_reports:
            try:
                logger.info(f"Running scheduled report: {report.name} (ID: {report.id})")
                
                # Create report run
                report_run = ReportRun.objects.create(
                    report=report,
                    status='running',
                    triggered_by=None  # System-triggered
                )
                
                # Generate report
                result = self.generator.generate_report(report, report_run)
                
                if result['success']:
                    report_run.mark_completed(
                        file_path=result['file_path'],
                        file_size=result['file_size'],
                        records_count=result['record_count']
                    )
                    
                    # Send email if configured
                    if report.email_recipients:
                        try:
                            self._send_report_email(report, report_run)
                            email_sent = True
                        except Exception as e:
                            logger.error(f"Failed to send email for report {report.id}: {str(e)}")
                            email_sent = False
                    else:
                        email_sent = None
                    
                    # Update next run time
                    report.next_run = self._calculate_next_run(report)
                    report.save(update_fields=['next_run'])
                    
                    results.append({
                        'report_id': report.id,
                        'report_name': report.name,
                        'status': 'success',
                        'message': 'Report generated successfully',
                        'email_sent': email_sent,
                        'file_size': result['file_size'],
                        'record_count': result['record_count']
                    })
                    
                else:
                    report_run.mark_failed(result['error'])
                    
                    results.append({
                        'report_id': report.id,
                        'report_name': report.name,
                        'status': 'failed',
                        'message': result['error']
                    })
                
            except Exception as e:
                logger.error(f"Unexpected error running scheduled report {report.id}: {str(e)}", exc_info=True)
                results.append({
                    'report_id': report.id,
                    'report_name': report.name,
                    'status': 'error',
                    'message': str(e)
                })
        
        logger.info(f"Completed scheduled reports run: {len(results)} reports processed")
        return results
    
    def _calculate_next_run(self, report: Report) -> Optional[datetime]:
        """Calculate next run time based on frequency"""
        return report.calculate_next_run()
    
    def _send_report_email(self, report: Report, report_run: ReportRun) -> None:
        """Send report via email"""
        subject = report.email_subject or f"Scheduled Report: {report.name}"
        body = report.email_body or f"""
Dear Recipient,

Please find attached the scheduled report: {report.name}

Report Details:
- Generated: {report_run.completed_at.strftime('%Y-%m-%d %H:%M:%S')}
- Records: {report_run.record_count or 'N/A'}
- File Size: {report_run.file_size_display}

Best regards,
Church Connect System
        """.strip()
        
        email = EmailMessage(
            subject=subject,
            body=body,
            from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@churchconnect.com'),
            to=report.email_recipients
        )
        
        # Attach report file
        if report_run.file_path and Path(report_run.file_path).exists():
            file_path = Path(report_run.file_path)
            email.attach_file(str(file_path))
        
        email.send()
        logger.info(f"Email sent successfully for report {report.id} to {len(report.email_recipients)} recipients")